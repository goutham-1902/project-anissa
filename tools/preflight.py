#!/usr/bin/env python3
from __future__ import annotations
from pathlib import Path
import json, os, sys, tempfile, zipfile
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
sys.path.insert(0,str(ROOT))
from project.environment import resolve_environment
ENVIRONMENT=resolve_environment(ROOT)
BRAIN=ENVIRONMENT.brain_path
SCHEMA=json.loads(ENVIRONMENT.schema_path.read_text())

def main():
    problems=[]
    if not BRAIN.exists(): problems.append('brain workbook missing')
    elif not zipfile.is_zipfile(BRAIN): problems.append('brain workbook is not a valid xlsx zip container')
    schema_version=SCHEMA.get('version')
    if not isinstance(schema_version,int) or schema_version < 1:
        problems.append('invalid schema version in schema file')
    required={
        'AGENTS.md':ROOT/'AGENTS.md',
        'START_HERE.md':ROOT/'START_HERE.md',
        'persona identity':ENVIRONMENT.private_persona_root/'anissa_identity.md',
        'verified profile':ENVIRONMENT.profile_root/'verified_profile.json',
        'canonical workbook':BRAIN,
    }
    if not (ENVIRONMENT.private_persona_root/'anissa_identity.md').exists():
        required['clean persona identity']=ENVIRONMENT.public_persona_root/'identity.md'
        required.pop('persona identity')
    for label,path in required.items():
        if not path.exists(): problems.append(f'missing {label}: {path}')
    settings=json.loads(ENVIRONMENT.runtime_settings_path.read_text())
    mode=settings.get('mode')
    authorized=settings.get('go_live_authorized')
    if mode not in {'SETUP','LIVE'}: problems.append(f'unsupported runtime mode: {mode!r}')
    if mode == 'SETUP' and authorized is not False:
        problems.append('SETUP mode requires go_live_authorized=false')
    if mode == 'LIVE' and authorized is not True:
        problems.append('LIVE mode requires go_live_authorized=true')
    if BRAIN.exists() and zipfile.is_zipfile(BRAIN):
        wb=load_workbook(BRAIN,read_only=False,data_only=False)
        missing=sorted(set(SCHEMA['sheets'])-set(wb.sheetnames))
        if missing: problems.append(f'missing workbook sheets: {missing}')
        control={row[0]:row[1] for row in wb['CONTROL'].iter_rows(min_row=2,values_only=True) if row[0]}
        if control.get('schema_version') != schema_version:
            problems.append('workbook CONTROL schema_version does not match schema file')
        if str(control.get('package_version') or '') != str(settings.get('package_version') or ''):
            problems.append('workbook CONTROL package_version does not match runtime settings')
        if control.get('setup_mode') != mode:
            problems.append('workbook CONTROL setup_mode does not match runtime mode')
        for sheet,expected in SCHEMA['sheets'].items():
            if not expected or sheet not in wb.sheetnames: continue
            actual=[c.value for c in wb[sheet][1]][:len(expected)]
            if actual != expected: problems.append(f'header mismatch on {sheet}')
        wb.close()
    for folder in [ENVIRONMENT.backup_root/'daily',ENVIRONMENT.backup_root/'weekly',ENVIRONMENT.lock_path.parent]:
        folder.mkdir(parents=True,exist_ok=True)
        try:
            fd,probe=tempfile.mkstemp(prefix='.preflight-',dir=folder)
            os.close(fd); Path(probe).unlink(missing_ok=True)
        except OSError as exc:
            problems.append(f'not writable: {folder} ({exc})')
    automations=Path.home()/'.codex'/'automations'
    active_anissa=[]
    if automations.exists():
        for path in automations.glob('*/automation.toml'):
            text=path.read_text(encoding='utf-8',errors='ignore')
            if 'anissa' in text.lower() and 'status = "ACTIVE"' in text:
                active_anissa.append(path.parent.name)
    bindings=settings.get('automation_bindings') or {}
    if mode == 'SETUP':
        if bindings: problems.append('SETUP mode must not contain automation bindings')
        if active_anissa: problems.append(f'Anissa automations active during SETUP: {active_anissa}')
    elif mode == 'LIVE':
        if not bindings: problems.append('LIVE mode requires automation bindings')
        expected={str(v) for v in bindings.values() if v}
        missing=sorted(expected-set(active_anissa))
        if missing: problems.append(f'bound Anissa automations are not active: {missing}')
    if problems:
        print('PREFLIGHT FAILED')
        for x in problems: print('-',x)
        raise SystemExit(1)
    print('PREFLIGHT OK')
    print(f'- package mode: {mode}')
    print('- brain workbook present')
    print(f'- schema version: {schema_version}')
    print('- workbook sheets and headers match schema')
    print('- lock and backup directories writable')
    print(f'- go-live authorization: {str(authorized).lower()}')
    if mode == 'SETUP':
        print('- no active Anissa automation found')
        print('- no automation activation performed')
    else:
        print(f'- active bound Anissa automations: {len(bindings)}')
if __name__=='__main__': main()
