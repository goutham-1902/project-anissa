from __future__ import annotations
from pathlib import Path
from datetime import datetime
import json, os, shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from .locks import exclusive_lock
from .task_state import validate_transition
from .ids import opportunity_id, task_id, application_id, scholarship_id, workload_event_id, weekly_audit_id
from .dedupe import equivalent_opportunity, equivalent_task, equivalent_application, equivalent_scholarship
from .workload import validate_workload_event, is_consequential_task_change, date_value
from .runtime import resolve_effective_mode
from project.environment import ProjectEnvironment, resolve_environment

ENVIRONMENT=resolve_environment(Path(__file__).resolve().parents[1])
ROOT=ENVIRONMENT.release_root
BRAIN=ENVIRONMENT.brain_path
LOCK=ENVIRONMENT.lock_path
EVENTS=ENVIRONMENT.events_path
ERRORS=ENVIRONMENT.errors_path


class WorkbookGateway:
    """The single mutation gateway for the canonical workbook.

    Production uses one package-wide lock/backups/events log. Tests or alternate workbooks
    automatically get isolated sibling lock/backups so validation cannot pollute live state.
    """
    def __init__(self, path: Path | None = None, *, environment: ProjectEnvironment | None = None,
                 allow_offline_migration: bool = False):
        self.environment=environment or ENVIRONMENT
        self.path=Path(path) if path is not None else self.environment.brain_path
        try:
            is_live=self.path.resolve() == self.environment.brain_path.resolve()
        except Exception:
            is_live=False
        self.is_canonical=is_live
        self.allow_offline_migration=allow_offline_migration
        if is_live:
            self.lock_path=self.environment.lock_path
            self.backup_root=self.environment.backup_root
            self.events_path=self.environment.events_path
            self.errors_path=self.environment.errors_path
        else:
            self.lock_path=self.path.with_suffix('.lock')
            self.backup_root=self.path.parent/'.anissa_test_backups'
            self.events_path=self.path.parent/'events.jsonl'
            self.errors_path=self.path.parent/'errors.jsonl'

    def _assert_mutation_allowed_locked(self):
        if not self.is_canonical or self.allow_offline_migration:
            return
        try:
            settings=json.loads(self.environment.runtime_settings_path.read_text(encoding='utf-8'))
        except (OSError, ValueError, TypeError) as exc:
            raise RuntimeError(
                f'Canonical mutation blocked: runtime settings unavailable ({exc.__class__.__name__}).'
            ) from exc
        gate=resolve_effective_mode(settings,self.read_control())
        if not gate['ok'] or gate['effective_mode'] != 'LIVE':
            reason='; '.join(gate['errors']) or 'effective mode is not LIVE'
            raise RuntimeError(f'Canonical mutation blocked: {reason}.')

    def _log(self, path: Path, event: dict):
        path.parent.mkdir(parents=True, exist_ok=True)
        event={'timestamp': datetime.now().isoformat(timespec='seconds'), **event}
        with path.open('a',encoding='utf-8') as f:
            f.write(json.dumps(event,ensure_ascii=False,default=str)+'\n')

    def _backup(self):
        stamp=datetime.now().strftime('%Y%m%d')
        dst=self.backup_root/'daily'/f'anissa_brain_{stamp}.xlsx'
        dst.parent.mkdir(parents=True, exist_ok=True)
        if not dst.exists():
            shutil.copy2(self.path,dst)
        lkg=self.backup_root/'anissa_brain_last_known_good.xlsx'
        lkg.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.path,lkg)
        iso_year, iso_week, _=datetime.now().isocalendar()
        weekly=self.backup_root/'weekly'/f'anissa_brain_{iso_year}-W{iso_week:02d}.xlsx'
        weekly.parent.mkdir(parents=True, exist_ok=True)
        if not weekly.exists():
            shutil.copy2(self.path,weekly)
        self._rotate_backups(self.backup_root/'daily','anissa_brain_*.xlsx',14)
        self._rotate_backups(self.backup_root/'weekly','anissa_brain_*.xlsx',12)

    @staticmethod
    def _rotate_backups(folder: Path, pattern: str, keep: int):
        backups=sorted(folder.glob(pattern),key=lambda p:p.stat().st_mtime,reverse=True)
        for old in backups[keep:]:
            old.unlink(missing_ok=True)

    def _atomic_save(self, wb):
        tmp=self.path.with_suffix('.tmp.xlsx')
        wb.save(tmp)
        os.replace(tmp,self.path)

    @staticmethod
    def _sheet_rows(ws) -> list[dict]:
        headers=[c.value for c in ws[1]]
        out=[]
        for vals in ws.iter_rows(min_row=2,values_only=True):
            if not any(v not in (None,'') for v in vals):
                continue
            out.append({headers[i]: vals[i] if i<len(vals) else None for i in range(len(headers))})
        return out

    def _upsert_locked(self, wb, sheet: str, id_column: str, row: dict) -> tuple[str,str]:
        ws=wb[sheet]
        headers=[c.value for c in ws[1]]
        if id_column not in headers:
            raise ValueError(f'{id_column} not found on {sheet}')
        if id_column not in row or not row[id_column]:
            raise ValueError(f'Missing {id_column}')
        idx=headers.index(id_column)+1
        target=None
        for r in range(2,ws.max_row+1):
            if str(ws.cell(r,idx).value or '') == str(row[id_column]):
                target=r; break
        if target is None:
            # Template sheets contain preformatted blank rows with validations. Reuse the
            # first genuinely empty row so inserts remain inside the existing Excel table.
            target=next(
                (
                    r for r in range(2,ws.max_row+1)
                    if not any(ws.cell(r,c).value not in (None,'') for c in range(1,len(headers)+1))
                ),
                ws.max_row+1,
            )
            action='insert'
            current={header:None for header in headers}
        else:
            action='update'
            current={header:ws.cell(target,i+1).value for i,header in enumerate(headers)}
        merged={**current,**row}
        if sheet == 'TASKS':
            if not str(merged.get('Definition of Done') or '').strip():
                raise ValueError('Meaningful tasks require a Definition of Done.')
            validate_transition(
                str(merged.get('Status') or 'Not Started'),
                evidence=str(merged.get('Evidence') or ''),
                blocker=str(merged.get('Blocker') or ''),
                unblock_action=str(merged.get('Unblock Action') or ''),
                meaningful=True,
            )
        for key,val in row.items():
            if key in headers:
                cell=ws.cell(target,headers.index(key)+1)
                if key in {'Start Date','End Date','Week Start','Week End'} and val not in (None,''):
                    val=date_value(val) or val
                    cell.number_format='yyyy-mm-dd'
                elif key in {'Created At','Updated At','Generated At'} and isinstance(val,datetime):
                    cell.number_format='yyyy-mm-dd hh:mm'
                cell.value=val
        for table in ws.tables.values():
            min_col,min_row,max_col,max_row=range_boundaries(table.ref)
            if min_row == 1 and target > max_row:
                table.ref=f'{get_column_letter(min_col)}1:{get_column_letter(max(max_col,len(headers)))}{target}'
        return action, str(row[id_column])

    def read_control(self) -> dict[str, object]:
        return self.read_bundle('CONTROL')['CONTROL']

    def rows(self, sheet: str) -> list[dict]:
        return self.read_bundle(sheet)[sheet]

    def read_bundle(self, *sheets: str) -> dict[str, object]:
        """Read related sheets from one immutable workbook generation."""
        if not sheets:
            raise ValueError('At least one sheet is required.')
        wb=load_workbook(self.path,read_only=True,data_only=False)
        try:
            result={}
            for sheet in sheets:
                if sheet == 'CONTROL':
                    result[sheet]={
                        str(row[0]):row[1]
                        for row in wb[sheet].iter_rows(min_row=2,values_only=True)
                        if row[0]
                    }
                else:
                    result[sheet]=self._sheet_rows(wb[sheet])
            return result
        finally:
            wb.close()

    def upsert(self, sheet: str, id_column: str, row: dict) -> str:
        """Exact-ID upsert. Prefer typed helpers below for cross-wording dedupe."""
        with exclusive_lock(self.lock_path):
            self._assert_mutation_allowed_locked()
            self._backup()
            wb=load_workbook(self.path)
            action, record_id=self._upsert_locked(wb,sheet,id_column,row)
            self._atomic_save(wb)
            wb.close()
            self._log(self.events_path,{'action':action,'sheet':sheet,'id':record_id})
            return action

    def _dedup_upsert(self, sheet: str, id_column: str, row: dict, eq_fn, id_factory) -> tuple[str,str]:
        with exclusive_lock(self.lock_path):
            self._assert_mutation_allowed_locked()
            self._backup()
            wb=load_workbook(self.path)
            ws=wb[sheet]
            existing=self._sheet_rows(ws)
            match=next((r for r in existing if eq_fn(r,row)),None)
            if match and match.get(id_column):
                row={**row,id_column:match[id_column]}
            elif not row.get(id_column):
                row={**row,id_column:id_factory(row)}
            action, record_id=self._upsert_locked(wb,sheet,id_column,row)
            self._atomic_save(wb)
            wb.close()
            self._log(self.events_path,{
                'action':action,'sheet':sheet,'id':record_id,
                'deduplicated_against_existing':bool(match)
            })
            return action, record_id

    def upsert_opportunity(self, row: dict) -> tuple[str,str]:
        def mk(r):
            return opportunity_id(str(r.get('Institution') or ''),str(r.get('Opportunity') or ''),str(r.get('Route') or ''),str(r.get('Official Source') or ''))
        return self._dedup_upsert('OPPORTUNITIES','Opportunity ID',row,equivalent_opportunity,mk)

    def upsert_task(self, row: dict) -> tuple[str,str]:
        def mk(r):
            return task_id(str(r.get('Application ID') or ''),str(r.get('Category') or ''),str(r.get('Task Title') or ''),str(r.get('Assigned Date') or ''))
        return self._dedup_upsert('TASKS','Task ID',row,equivalent_task,mk)

    def commit_daily_plan(self, tasks: list[dict]) -> list[dict]:
        """Atomically deduplicate, persist and read back one reported task bundle."""
        if not tasks:
            raise ValueError('Daily plan requires at least one task.')
        with exclusive_lock(self.lock_path):
            self._assert_mutation_allowed_locked()
            self._backup()
            wb=load_workbook(self.path)
            results=[]
            try:
                ws=wb['TASKS']
                existing=self._sheet_rows(ws)
                for raw in tasks:
                    row=dict(raw)
                    match=next((item for item in existing if equivalent_task(item,row)),None)
                    if match and match.get('Task ID'):
                        row={**row,'Task ID':match['Task ID']}
                    elif not row.get('Task ID'):
                        row={**row,'Task ID':task_id(
                            str(row.get('Application ID') or ''),str(row.get('Category') or ''),
                            str(row.get('Task Title') or ''),str(row.get('Assigned Date') or ''),
                        )}
                    action,record_id=self._upsert_locked(wb,'TASKS','Task ID',row)
                    persisted=next(item for item in self._sheet_rows(ws) if str(item.get('Task ID'))==record_id)
                    results.append({
                        'action':action,'task_id':record_id,'title':persisted.get('Task Title'),
                        'due_date':persisted.get('Due Date'),'status':persisted.get('Status'),
                    })
                    existing=[item for item in existing if str(item.get('Task ID'))!=record_id]+[persisted]
                self._atomic_save(wb)
            finally:
                wb.close()
        committed={str(row.get('Task ID')) for row in self.rows('TASKS')}
        missing=[row['task_id'] for row in results if row['task_id'] not in committed]
        if missing:
            raise RuntimeError(f'Daily plan read-back failed: {missing}')
        self._log(self.events_path,{'action':'commit_daily_plan','tasks':results})
        return results

    def upsert_application(self, row: dict) -> tuple[str,str]:
        def mk(r):
            return application_id(str(r.get('Target ID') or ''),str(r.get('Institution') or ''),str(r.get('Programme / Position') or ''),str(r.get('Cycle') or ''))
        return self._dedup_upsert('APPLICATIONS','Application ID',row,equivalent_application,mk)

    def upsert_scholarship(self, row: dict) -> tuple[str,str]:
        def mk(r):
            return scholarship_id(str(r.get('Target') or ''),str(r.get('Sponsor') or ''),str(r.get('Scholarship') or ''),str(r.get('Deadline') or ''))
        return self._dedup_upsert('SCHOLARSHIPS','Scholarship ID',row,equivalent_scholarship,mk)

    def upsert_workload_event(self, row: dict) -> tuple[str,str]:
        validate_workload_event(row)
        if not row.get('Event ID'):
            row={**row,'Event ID':workload_event_id(
                str(row.get('Event Type') or ''),str(row.get('Title') or ''),
                row.get('Start Date') or '',row.get('End Date') or '',
            )}
        now=datetime.now()
        row={**row,'Updated At':now}
        existing=next((r for r in self.rows('WORKLOAD_EVENTS') if r.get('Event ID')==row['Event ID']),None)
        if not existing:
            row={'Created At':now,**row}
        return self.upsert('WORKLOAD_EVENTS','Event ID',row),str(row['Event ID'])

    def apply_workload_replan(self, event: dict, task_changes: list[dict]) -> dict:
        """Apply one approved workload event and all task changes in one transaction."""
        validate_workload_event(event)
        event=dict(event)
        event_id_value=str(event.get('Event ID') or workload_event_id(
            str(event.get('Event Type') or ''),str(event.get('Title') or ''),
            event.get('Start Date') or '',event.get('End Date') or '',
        ))
        event['Event ID']=event_id_value
        now=datetime.now()
        with exclusive_lock(self.lock_path):
            self._assert_mutation_allowed_locked()
            self._backup()
            wb=load_workbook(self.path)
            try:
                task_rows=self._sheet_rows(wb['TASKS'])
                tasks_by_id={str(row.get('Task ID')):row for row in task_rows if row.get('Task ID')}
                consequential=[]
                prepared=[]
                for raw in task_changes:
                    change={**raw,'Workload Event ID':raw.get('Workload Event ID') or event_id_value}
                    current=tasks_by_id.get(str(change.get('Task ID') or ''))
                    if current and is_consequential_task_change(current,change):
                        consequential.append(str(current['Task ID']))
                    if current:
                        change={'Definition of Done':current.get('Definition of Done'),**change}
                    elif not change.get('Task ID'):
                        match=next((row for row in task_rows if equivalent_task(row,change)),None)
                        if match:
                            change={**change,'Task ID':match['Task ID']}
                        else:
                            change={**change,'Task ID':task_id(
                                str(change.get('Application ID') or ''),str(change.get('Category') or ''),
                                str(change.get('Task Title') or ''),str(change.get('Assigned Date') or ''),
                            )}
                    prepared.append(change)
                if consequential and str(event.get('Consequential Approval') or '').upper() != 'YES':
                    raise ValueError(f'Consequential task changes require explicit approval: {consequential}')
                existing_event=next((r for r in self._sheet_rows(wb['WORKLOAD_EVENTS']) if r.get('Event ID')==event_id_value),None)
                event={
                    **event,
                    'Affected Task IDs':'; '.join(str(row.get('Task ID')) for row in prepared if row.get('Task ID')),
                    'Created At':(existing_event or {}).get('Created At') or now,
                    'Updated At':now,
                }
                event_action,_=self._upsert_locked(wb,'WORKLOAD_EVENTS','Event ID',event)
                task_results=[]
                for change in prepared:
                    action,record_id=self._upsert_locked(wb,'TASKS','Task ID',change)
                    task_results.append({'action':action,'task_id':record_id})
                self._atomic_save(wb)
            finally:
                wb.close()
            self._log(self.events_path,{
                'action':'apply_workload_replan','event_id':event_id_value,
                'event_action':event_action,'tasks':task_results,
            })
            return {'event_id':event_id_value,'event_action':event_action,'tasks':task_results}

    def record_reminder(self, task_id_value: str, *, at: datetime | None = None):
        with exclusive_lock(self.lock_path):
            self._assert_mutation_allowed_locked()
            self._backup()
            wb=load_workbook(self.path); ws=wb['TASKS']
            try:
                headers=[c.value for c in ws[1]]
                id_col=headers.index('Task ID')+1
                rownum=next((r for r in range(2,ws.max_row+1) if str(ws.cell(r,id_col).value or '')==task_id_value),None)
                if rownum is None:
                    raise KeyError(task_id_value)
                status=str(ws.cell(rownum,headers.index('Status')+1).value or '')
                if status != 'Not Started':
                    raise ValueError(f'Reminder not allowed for task in {status!r}.')
                count_cell=ws.cell(rownum,headers.index('Reminder Count')+1)
                count_cell.value=int(count_cell.value or 0)+1
                ws.cell(rownum,headers.index('Last Reminder')+1).value=at or datetime.now()
                self._atomic_save(wb)
            finally:
                wb.close()
            self._log(self.events_path,{'action':'task_reminder','task_id':task_id_value})

    def record_weekly_audit(self, row: dict) -> tuple[str,str]:
        week_start=row.get('Week Start')
        if not week_start:
            raise ValueError('Weekly audits require Week Start.')
        row=dict(row)
        row['Audit ID']=row.get('Audit ID') or weekly_audit_id(week_start)
        row['Generated At']=row.get('Generated At') or datetime.now()
        return self.upsert('WEEKLY_AUDITS','Audit ID',row),str(row['Audit ID'])

    def set_control(self, key: str, value: object):
        with exclusive_lock(self.lock_path):
            self._assert_mutation_allowed_locked()
            self._backup()
            wb=load_workbook(self.path); ws=wb['CONTROL']
            found=False
            for r in range(2,ws.max_row+1):
                if ws.cell(r,1).value==key:
                    ws.cell(r,2).value=value; found=True; break
            if not found:
                ws.append([key,value,'Added by runtime'])
            self._atomic_save(wb); wb.close()
            self._log(self.events_path,{'action':'set_control','key':key,'value':value})

    def set_task_status(self, task_id_value: str, status: str, *, evidence: str='', blocker: str='', unblock_action: str='', actual_minutes: int | None=None):
        with exclusive_lock(self.lock_path):
            self._assert_mutation_allowed_locked()
            self._backup()
            wb=load_workbook(self.path); ws=wb['TASKS']
            headers=[c.value for c in ws[1]]
            id_col=headers.index('Task ID')+1
            rownum=None
            for r in range(2,ws.max_row+1):
                if str(ws.cell(r,id_col).value or '')==task_id_value:
                    rownum=r; break
            if rownum is None:
                raise KeyError(task_id_value)
            meaningful=True
            validate_transition(status,evidence=evidence,blocker=blocker,unblock_action=unblock_action,meaningful=meaningful)
            def setv(name,val):
                ws.cell(rownum,headers.index(name)+1).value=val
            now=datetime.now()
            setv('Status',status)
            if status=='Started' and not ws.cell(rownum,headers.index('Started At')+1).value:
                setv('Started At',now)
            if status=='Done':
                setv('Completed At',now); setv('Evidence',evidence)
                if actual_minutes is not None and 'Actual Minutes' in headers:
                    setv('Actual Minutes',actual_minutes)
                if 'Remaining Minutes' in headers:
                    setv('Remaining Minutes',0)
            if status=='Blocked':
                setv('Blocked Since',now); setv('Blocker',blocker); setv('Unblock Action',unblock_action)
            self._atomic_save(wb); wb.close()
            self._log(self.events_path,{'action':'task_status','task_id':task_id_value,'status':status})
