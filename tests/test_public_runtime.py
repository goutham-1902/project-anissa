from hashlib import sha256
import json
import os
from pathlib import Path
import unittest

from openpyxl import load_workbook

from anissa.core import AnissaCore
from logic.workbook_io import WorkbookGateway
from project.environment import RELEASE_ROOT, resolve_environment
from project.governance import Governance
from project.telemetry_contract import read_publication
from worker1.src.dashboard_server import dashboard_build_id, dashboard_health


class PublicRuntimeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.environment = resolve_environment(RELEASE_ROOT)

    def test_synthetic_instance_is_setup_and_schema_complete(self):
        settings = json.loads(self.environment.runtime_settings_path.read_text())
        self.assertEqual(settings["mode"], "SETUP")
        self.assertFalse(settings["go_live_authorized"])
        workbook = load_workbook(self.environment.brain_path, read_only=True)
        schema = json.loads(self.environment.schema_path.read_text())
        self.assertEqual(set(workbook.sheetnames), set(schema["sheets"]))
        workbook.close()

    def test_core_is_blocked_before_go_live_and_mutation_fails_closed(self):
        gateway = WorkbookGateway(environment=self.environment)
        before = sha256(self.environment.brain_path.read_bytes()).digest()
        snapshot = AnissaCore(self.environment, gateway=gateway).snapshot("status")
        self.assertTrue(snapshot["blocked"])
        with self.assertRaisesRegex(RuntimeError, "mutation blocked"):
            gateway.set_control("probe", "blocked")
        self.assertEqual(before, sha256(self.environment.brain_path.read_bytes()).digest())

    def test_empty_telemetry_is_checksum_coherent_not_false_activity(self):
        publication = read_publication(
            self.environment.telemetry_root / "worklog.csv",
            self.environment.telemetry_root / "status.json",
        )
        self.assertEqual(publication.rows, [])
        self.assertEqual(publication.status["state"], "SETUP")

    def test_worker_source_cannot_import_campaign_gateway(self):
        worker = RELEASE_ROOT / "worker1" / "src"
        text = "\n".join(path.read_text() for path in worker.glob("*.py"))
        self.assertNotIn("WorkbookGateway", text)
        self.assertNotIn("anissa.agendas", text)

    def test_dashboard_includes_bounded_weekly_history_without_private_assets(self):
        index = (RELEASE_ROOT / "worker1" / "dashboard" / "index.html").read_text()
        app = (RELEASE_ROOT / "worker1" / "dashboard" / "app.js").read_text()
        self.assertIn('id="weeklyHistory"', index)
        self.assertIn("weekly_history", app)
        self.assertNotIn("anissa-verdict.png", index)

    def test_dashboard_health_identifies_the_current_release_build(self):
        build_id = dashboard_build_id(RELEASE_ROOT)
        health = dashboard_health(build_id)
        self.assertEqual(health["service"], "a2a-dashboard")
        self.assertEqual(health["build_id"], build_id)
        self.assertEqual(len(build_id), 16)

    def test_private_instance_and_clean_presentation_are_separate(self):
        self.assertFalse((RELEASE_ROOT / "brain").exists())
        self.assertFalse((RELEASE_ROOT / "profile").exists())
        self.assertFalse((RELEASE_ROOT / "persona").exists())
        self.assertTrue(self.environment.public_persona_root.is_dir())
        self.assertTrue(self.environment.instance_root.is_relative_to(
            Path(os.environ["PROJECT_ANISSA_INSTANCE"]).resolve()
        ))

    def test_maintainer_scope_defers_before_execution(self):
        governance = Governance(self.environment)
        governance.initialize_ledgers("2.5.0-dev.8")
        decision = governance.evaluate_scope(
            "SOLDIERS_MAINTAINER",
            ["worker1/src/sync.py", "project/projections.py"],
        )
        self.assertFalse(decision["accepted"])
        self.assertEqual(decision["defer_to"], "GENERAL")

        route = governance.publication_route(
            "SOLDIERS_MAINTAINER", ["worker1/src/sync.py"]
        )
        self.assertTrue(route["accepted"])
        self.assertEqual(route["publisher"], "GENERAL")
        self.assertEqual(route["action"], "HANDOFF_VERIFIED_CHANGE_TO_GENERAL")


if __name__ == "__main__":
    unittest.main()
